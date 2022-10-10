import shutil
import datetime
import decimal
import openpyxl
import os

import pprint

TEMPLATE_FILENAME = 'sys\\template.xlsx'
XLSX_FOLDER = 'EXCEL_XLSX'
LOG_FILENAME = 'main.log'
PP = pprint.PrettyPrinter()


class DataNotFound(Exception):
    """Не найдена информация с указанными годом и кодом МСЗ"""


class MSZCountNotInt(Exception):
    """Количество назначений не целое число"""


class MSZSumNotDecimal(Exception):
    """Сумма назначений не число"""


def add_msg_in_log_file(msg: str, filename: str = LOG_FILENAME, mode: str = 'a', encoding: str = 'utf-8'):
    with open(filename, mode, encoding=encoding) as out_file:
        out_file.write(f'{msg}\n')


def create_log_file(time_stamp: datetime = datetime.datetime.now()):
    add_msg_in_log_file(f'Начало работы: {time_stamp.strftime("%Y.%m.%d %H:%M")}\n', mode='w')


def create_xlsx_file_from_template(template_fn: str = TEMPLATE_FILENAME,
                                   time_stamp: datetime = datetime.datetime.now()):
    shutil.copyfile(template_fn, f'COUNT_{time_stamp.strftime("%d_%m_%Y_%H%M")}.xlsx')
    return f'COUNT_{time_stamp.strftime("%d_%m_%Y_%H%M")}.xlsx'


def find_data_in_workbook(wb: openpyxl.Workbook, year: str, msz_code: str, mode: str = 'r',
                          stat_dict: dict = None) -> (int, decimal):
    ws = wb.active
    for row in ws.iter_rows():
        if str(row[2].value) == year and str(row[3].value).zfill(4) == msz_code:
            if mode == 'r':
                try:
                    if row[5].value is None or str(row[5].value) == '0':
                        return 0, decimal.Decimal(0)
                    msz_count = int(row[5].value)
                except Exception:
                    raise MSZCountNotInt
                try:
                    if row[6].value is None or str(row[6].value) == '0':
                        return 0, decimal.Decimal(0)
                    msz_sum = decimal.Decimal(row[6].value)
                except Exception:
                    raise MSZSumNotDecimal

                return msz_count, msz_sum
            elif mode == 'w':
                row[5].value = stat_dict[year][msz_code]['count']
                row[6].value = float(stat_dict[year][msz_code]['sum'])
                return None

    raise DataNotFound


def create_filenames_list(folder_name: str = XLSX_FOLDER, ext_filter: str = '.xlsx') -> list:
    filenames_list = []

    for root, dirs, files in os.walk(folder_name):
        for file in files:
            if file.endswith(ext_filter) and not file.startswith('~'):
                filenames_list.append(os.path.join(root, file))

    return filenames_list


def create_dict_from_workbook(wb: openpyxl.Workbook, start_row: int = 8):
    ws = wb.active

    result_dict = dict()
    row_i = start_row

    try:
        is_data_row = int(ws[f'A{row_i}'].value) == (row_i - start_row + 1)
    except (Exception, ):
        return result_dict

    while is_data_row:

        lmsz_year = str(ws[f'C{row_i}'].value)
        lmsz_code = str(ws[f'D{row_i}'].value).zfill(4)

        if lmsz_year not in result_dict.keys():
            result_dict[lmsz_year] = {lmsz_code: {'count': 0, 'sum': decimal.Decimal(0)}}
        else:
            result_dict[lmsz_year][lmsz_code] = {'count': 0, 'sum': decimal.Decimal(0)}
        row_i += 1
        try:
            is_data_row = int(ws[f'A{row_i}'].value) == (row_i - start_row + 1)
        except (Exception,):
            return result_dict

    return result_dict


def main():
    create_log_file()

    xlsx_stat_dict = create_dict_from_workbook(openpyxl.load_workbook(filename=TEMPLATE_FILENAME, data_only=True))

    xlsx_filelist = create_filenames_list()

    for xlsx_filename in xlsx_filelist:
        wb = openpyxl.load_workbook(filename=xlsx_filename, data_only=True)
        add_msg_in_log_file(f'Обработка файла: {xlsx_filename}')
        for lmsz_year in xlsx_stat_dict.keys():
            for lmsz_code in xlsx_stat_dict[lmsz_year].keys():
                try:
                    lmsz_count, lmsz_sum = find_data_in_workbook(wb, lmsz_year, lmsz_code)
                    xlsx_stat_dict[lmsz_year][lmsz_code]['count'] += lmsz_count
                    xlsx_stat_dict[lmsz_year][lmsz_code]['sum'] += lmsz_sum
                except Exception as e:
                    add_msg_in_log_file(f'{lmsz_year}:{lmsz_code} - {e}')

    PP.pprint(xlsx_stat_dict)
    xlsx_stat_filename = create_xlsx_file_from_template()
    wb = openpyxl.load_workbook(filename=xlsx_stat_filename, data_only=True)
    for lmsz_year in xlsx_stat_dict.keys():
        for lmsz_code in xlsx_stat_dict[lmsz_year].keys():
            find_data_in_workbook(wb, mode='w', year=lmsz_year, msz_code=lmsz_code, stat_dict=xlsx_stat_dict)

    wb.save(xlsx_stat_filename)


if __name__ == '__main__':
    main()
